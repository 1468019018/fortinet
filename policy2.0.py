import re, os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import tkinter as tk
from tkinter import filedialog
from pathlib import Path

input_file_path = ""
output_file_path = ""


# 弹窗获取文件路径
def select_file_and_get_path():
    # 创建一个隐藏的主窗口
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    # 弹出文件选择对话框
    file_path = filedialog.askopenfilename(
        title="请选择配置文件",
        filetypes=[("配置文件", "*.conf"), ("文本文件", "*.txt")]  # 可以根据需要调整文件类型
    )
    file_path = Path(file_path)

    file_name_with_extension = os.path.basename(file_path)
    input_filename = os.path.splitext(file_name_with_extension)[0]

    # 返回选择的文件路径
    return file_path, input_filename


# 弹窗选择导出路径
def select_folder_and_get_path(input_filename):
    # 创建一个隐藏的主窗口
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    # 弹出文件夹选择对话框
    folder_path = filedialog.askdirectory(
        title="请选择导出文件路径！"
    )

    # 检查是否选择了文件夹
    if not folder_path:
        print("未选择文件夹，进程将中止。")
        sys.exit(1)  # 使用 sys.exit() 中止进程，1 表示异常退出

    # 构建完整的文件路径
    folder_path = Path(folder_path)
    file_path = str(folder_path) + '\\' + input_filename + "-用户策略" + ".xlsx"

    # 判断原来是否存在相应文件，如果有，删除
    if os.path.exists(file_path):
        if os.path.isfile(file_path):
            os.remove(file_path)
            print(f"文件 '{file_path}' 已覆盖。")

    # 返回选择的文件路径
    return file_path


# 通过配置文件截取配置，title输入需要截取的段落
def extract_config_content(file_path, title, title2=""):
    """
    从给定的文本中提取 config user local 和 行首的 end 之间的内容。

    参数:
    text (str): 包含配置块的文本。

    返回:
    str: 提取的内容，去除空行和首尾空格。
    """
    # 使用正则表达式提取 config user local 和 行首的 end 之间的内容
    with open(file_path, 'r', encoding='utf-8') as file:
        text = file.read()
    output_text = ""
    for txt in [title, title2]:
        pattern = txt + r"\s+(.*?)(?=\n^end)"
        match = re.search(pattern, text, re.DOTALL | re.MULTILINE)
        if match:
            content = match.group(1).strip()
            # 由于 .*? 是非贪婪匹配，并且我们使用了 re.DOTALL 来匹配换行符，
            # 但我们可能需要进一步处理内容以去除不必要的缩进或空行
            lines = content.splitlines()
            filtered_lines = [line.strip() for line in lines if line.strip()]  # 去除空行和首尾空格
            output_text += "\n".join(filtered_lines)
        else:
            return "No match found."
    return output_text


# 提取用户表
def get_user_list(text):
    user_list = []
    pattern = re.compile(r'edit\s+"(.*?)"')
    user_list = pattern.findall(text)
    return user_list
# 提取列表
def extract_dstaddr_lines(text, pattern=r'set\s+dstaddr\s+(.*)'):
    # 用于存储所有提取的dstaddr元素
    dstaddr_elements = []

    # 按行分割文本
    lines = text.strip().split('\n')

    # 遍历每一行
    for line in lines:
        # 使用正则表达式查找set dstaddr后的内容
        match = re.search(pattern, line)
        if match:
            # 提取匹配到的内容，并按引号分割获取每个元素
            elements = re.findall(r'"(.*?)"', match.group(1))
            dstaddr_elements.extend(elements)  # 将当前行的元素添加到总列表中

    # 去重并转换为列表
    unique_dstaddr_elements = list(set(dstaddr_elements))

    return unique_dstaddr_elements

# 提取用户-用户组表
def user_group(text):
    # 用于存储结果的字典
    groups_user_dict = {}

    # 正则表达式模式
    edit_pattern = re.compile(r'edit\s+"(.*?)"')
    set_member_pattern = re.compile(r'set\s+member\s+(.*)')  # 捕获 set member 后的所有内容

    for line in text.strip().split('\n'):
        line = line.strip()
        # 匹配 edit 行
        edit_match = edit_pattern.match(line)

    # 临时变量，用于存储当前处理的组名和成员
    current_group = None
    members = []

    # 逐行解析文本
    for line in text.strip().split('\n'):
        line = line.strip()

        # 匹配 edit 行
        edit_match = edit_pattern.match(line)
        if edit_match:
            # 如果之前有组正在处理，则保存该组
            if current_group:
                groups_user_dict[current_group] = members

            # 开始处理新组
            current_group = edit_match.group(1)
            members = []  # 重置成员列表

        # 匹配 set member 行
        elif current_group and (set_member_match := set_member_pattern.match(line)):
            # 提取 set member 后的所有内容
            member_content = set_member_match.group(1)

            # 使用正则表达式提取所有引号包围的成员
            members = re.findall(r'"(.*?)"', member_content)

    # 保存最后一个组（如果有的话）
    if current_group:
        groups_user_dict[current_group] = members
    return groups_user_dict

# 获取IP组，IP列表
def ipgroup_ip(text):
    config_dict = {}
    current_key = None
    current_value_lines = []
    start_end_list = []
    # 按行分割文本
    lines = text.strip().split('\n')
    # print(lines)
    for line in lines:
        line = line.strip()

        # 检查是否是 edit 行
        edit_match = re.match(r'edit\s+"(.*?)"', line)
        if edit_match:
            current_key = edit_match.group(1)

        elif current_key and (line.startswith('set member')):
            start_index = line.find('set member ') + len('set member ')
            members_part = line[start_index:]
            config_dict[current_key] = re.findall(r'"(.*?)"', members_part)
        elif current_key and (line.startswith('set subnet')):
            start_index = line.find('set subnet ') + len('set subnet ')
            members_part = line[start_index:]
            config_dict[current_key] =["subnet:"+members_part]

        elif current_key and (line.startswith('set start-ip')):
            line = line[4:]
            start_end_list.append(line)
        elif current_key and (line.startswith('set end-ip')):
            line = line[4:]
            start_end_list.append(line)
            config_dict[current_key] = start_end_list
            start_end_list = []
        # print(config_dict)
    return config_dict


# 获取用户、用户组、目的IP、服务字典
def parse_groups_and_users(text):
    # 初始化结果字典
    groups_dict = {}
    name_dict = {}
    users_dict = {}
    dstaddr_dict = {}
    service_dict = {}
    status_dict = {}
    # 默认表名（根据 set name 提取）
    table_name = None

    # 正则表达式模式
    # 表名
    set_id_pattern = re.compile(r'edit\s+(.*)')
    # policy name
    set_name_pattern = re.compile(r'set\s+name\s+(.*)')
    # 用户组名
    set_groups_pattern = re.compile(r'set\s+groups\s+(.*)')  # 捕获 set groups 后的内容
    # 用户名
    set_users_pattern = re.compile(r'set\s+users\s+(.*)')  # 捕获 set users 后的内容
    # 目的IP名
    set_dstaddr_pattern = re.compile(r'set\s+dstaddr\s+(.*)')  # 捕获 set dstaddr 后的内容
    # 目的服务
    set_service_pattern = re.compile(r'set\s+service\s+(.*)')  # 捕获 set service 后的内容
    # 是否有效
    set_status_pattern = re.compile(r'set\s+status\s+(.*)')  # 捕获 set status 后的内容

    # 逐行解析文本
    for line in text.strip().split('\n'):
        line = line.strip()

        # 匹配 set name 行
        set_id_match = set_id_pattern.match(line)
        if set_id_match:
            table_name = set_id_match.group(1)  # 提取表名
        # 匹配name 行
        elif table_name and 'set name' in line:
            # 提取 set name 后的所有内容，并分割成列表
            name_match = set_name_pattern.search(line)
            name_content = name_match.group(1)
            name_list = [name_content]
            name_dict[table_name] = name_list

        # 匹配 set status 行
        elif table_name and 'set status' in line:
            # 提取 set status 后的所有内容，并分割成列表
            status_match = set_status_pattern.search(line)
            status_content = status_match.group(1)
            status_list = [status_content]
            status_dict[table_name] = status_list
            # if status_match:
            #     status_content = status_match.group(1)
            #     status_list = [status.strip('\b') for status in re.findall(r'"(.*?)"', status_content)]
            #     status_dict[table_name] = status_list

        # 匹配 set groups 行
        elif table_name and 'set groups' in line:
            # 提取 set groups 后的所有内容，并分割成列表
            groups_match = set_groups_pattern.search(line)
            if groups_match:
                groups_content = groups_match.group(1)
                # 使用 re.findall 来处理可能的多个元素，包括引号中的空格等特殊情况
                groups_list = [group.strip('"') for group in re.findall(r'"(.*?)"', groups_content)]
                groups_dict[table_name] = groups_list

        # 匹配 set users 行
        elif table_name and 'set users' in line:
            # 提取 set users 后的所有内容，并分割成列表
            users_match = set_users_pattern.search(line)
            if users_match:
                users_content = users_match.group(1)
                users_list = [user.strip('"') for user in re.findall(r'"(.*?)"', users_content)]
                users_dict[table_name] = users_list
        # 匹配 set dstaddr 行
        elif table_name and 'set dstaddr' in line:
            # 提取 set dstaddr 后的所有内容，并分割成列表
            dstaddr_match = set_dstaddr_pattern.search(line)
            if dstaddr_match:
                dstaddr_content = dstaddr_match.group(1)
                dstaddr_list = [dstaddr.strip('"') for dstaddr in re.findall(r'"(.*?)"', dstaddr_content)]
                dstaddr_dict[table_name] = dstaddr_list

        # 匹配 set service 行
        elif table_name and 'set service' in line:
            # 提取 set service 后的所有内容，并分割成列表
            service_match = set_service_pattern.search(line)
            if service_match:
                service_content = service_match.group(1)
                service_list = [service.strip('"') for service in re.findall(r'"(.*?)"', service_content)]
                service_dict[table_name] = service_list


    return groups_dict, users_dict, dstaddr_dict, service_dict, status_dict, name_dict


# 构建字典反向映射表
def rever_dict(group):
    # 构建用户到用户组的反向映射
    rever_groups = {}
    for group, users in group_user_map.items():
        for user in users:
            if user not in rever_groups:
                rever_groups[user] = []
            rever_groups[user].append(group)
    return rever_groups


# 写入excel，文件路径，sheet名称，字典，A列，B列
def edit_excel_sheet(file_path, sheet_name, groups, A_name="列1", B_name="列2"):
    if not os.path.exists(file_path):
        wb = Workbook()  # 创建新的工作簿
        ws = wb.active  # 获取默认的 sheet
        ws.title = sheet_name  # 默认将第一个 sheet 命名为指定的 sheet_name
        # print(f"文件 {file_path} 不存在，已创建新文件。")
    else:
        # 加载现有的工作簿
        wb = load_workbook(file_path)
        # 检查 sheet 是否存在，如果不存在则创建
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            print(f"Sheet '{sheet_name}' 不存在，已创建新 sheet。")

        else:
            ws = wb[sheet_name]
    if type(groups) == dict:
        # 清空现有的内容（可选，如果你希望保留现有内容，可以注释掉这一行）
        ws.delete_rows(1, ws.max_row)

        # 写入表头
        ws.append([A_name, B_name])
        # 2025/4/16 16：12 不在这里合并单元格，最后统一合并单元格子，若需要修改，把下面的！= 后面的111去掉

        for table_name, users in groups.items():
            for user in users:
                ws.append([table_name, user])

    elif type(groups) == list:
        # 清空现有的内容（可选，如果你希望保留现有内容，可以注释掉这一行）
        ws.delete_rows(1, ws.max_row)

        # 写入表头
        ws.append([A_name, B_name])
        for user in groups:
            print(user)
            ws.append([user])
    # 保存工作簿
    wb.save(file_path)

# 通过比对用户名、策略，填写目标信息
def fill_excel_column(file_path, map, title, source_col="A", target_col="C", start_row=2):
    """
    遍历Excel指定列，根据字典映射填充目标列

    参数:
    file_path：文件路径
    source_col (str): 源列字母（如'A'）
    target_col (str): 目标列字母（如'C'）
    map (dict): 字典映射关系 {键: 值}
    start_row (int): 起始行号（默认从第1行开始）
    """
    wb = load_workbook(file_path)
    ws = wb.active
    # 将列字母转换为数字索引
    source_col_num = column_index_from_string(source_col)
    target_col_num = column_index_from_string(target_col)

    current_row = start_row
    #设置标题
    ws.cell(row=1, column=target_col_num).value = title
    while True:
        # 获取当前单元格的值
        cell = ws.cell(row=current_row, column=source_col_num)
        cell_value = cell.value

        # 遇到空单元格时停止循环
        if cell_value is None:
            break
        # 查询是否存在多个值
        elif '\n' in cell_value:
            cell_value_list = cell_value.split('\n')
            for cell_value in cell_value_list:
                # cell_value = [cell_value]
                target_value = map.get(cell_value)
                write_list_to_excel(ws, target_value, row=current_row, col=target_col_num)
        else:
            # 查找字典中的对应值
            target_value = map.get(cell_value)
            # 如果找到映射值，则写入目标列
            if target_value is not None:
                # target_cell = ws.cell(row=current_row, column=target_col_num)
                # print(target_value)
                write_list_to_excel(ws, target_value, row=current_row, col=target_col_num)
                # target_cell.value = target_value


        current_row += 1
    wb.save(file_path)

# # 匹配excel的策略，填充IP组，服务组
# def process_excel(file_path, user_policy_map, policy_att_map, c_column_title, target_column='C', match_column='B'):
#     # 打开 Excel 文件
#     wb = load_workbook(file_path)
#
#     # 获取 Sheet1 和 Sheet2
#     sheet1 = wb[user_policy_map]
#     sheet2 = wb[policy_att_map]
#
#     # 确定目标列的索引
#     target_col_index = ord(target_column.upper()) - ord('A') + 1
#     target_col_letter = target_column.upper()
#     # 设置 Sheet1 的目标列标题
#     sheet1[f'{target_col_letter}1'] = c_column_title
#
#     # 设置目标列的自动换行
#     for row in range(1, sheet1.max_row + 1):
#         cell = sheet1[f'{target_col_letter}{row}']
#         cell.alignment = Alignment(wrapText=True)
#
#     # 遍历 Sheet1 的 B 列，获取数据用于查询;(默认B列，可以入参match_column修改)
#     for row in range(2, sheet1.max_row + 1):  # 假设第一行是标题行，从第二行开始
#         key_value = sheet1[f'{match_column}{row}'].value
#         if key_value is not None:
#             # 在 Sheet2 中查找对应的 A 列值，并收集所有匹配的 B 列值
#             matching_values = []
#             for sheet2_row in range(2, sheet2.max_row + 1):  # 假设第一行是标题行，从第二行开始
#                 if sheet2[f'A{sheet2_row}'].value == key_value:
#                     # 获取对应的 B 列值
#                     matching_values.append(sheet2[f'B{sheet2_row}'].value)
#
#             # 将所有匹配的值用换行符连接，并填入 Sheet1 的目标列
#             sheet1[f'{target_col_letter}{row}'] = '\n'.join(
#                 str(value) for value in matching_values if value is not None)
#         else:
#             # 如果 B 列值为空，目标列也设为空
#             sheet1[f'{target_col_letter}{row}'] = ''
#
#     # 保存工作簿
#     wb.save(file_path)
#

# 根据用户名，匹配对应策略

def find_user_policies(user_list, user_policy_map, group_policy_map, group_user_map):
    # 用于存储结果的字典
    user_policy_dict = {user: set() for user in user_list}

    # 1. 遍历用户列表，针对每个用户先查询一次【用户-策略表】
    for policy_str, users in user_policy_map.items():
        # policy = policy_str.strip('"')  # 去掉引号以匹配逻辑（如果需要）
        policy = policy_str
        for user in users:
            if user in user_policy_dict:
                user_policy_dict[user].add(policy)

    # 构建用户到用户组的反向映射
    user_to_groups = {}
    for group, users in group_user_map.items():
        for user in users:
            if user not in user_to_groups:
                user_to_groups[user] = []
            user_to_groups[user].append(group)
    # 构建用户组到策略的反向映射
    groups_to_policy = {}
    for policy, groups in group_policy_map.items():
        for group in groups:
            if group not in groups_to_policy:
                groups_to_policy[group] = []
            groups_to_policy[group].append(policy)

    # 2. 遍历用户列表，继续通过【用户-用户组表】（反向映射）和【用户组-策略表】查找策略
    for user in user_list:
        if user in user_to_groups:
            user_groups = user_to_groups[user]
            for group in user_groups:
                if group in groups_to_policy:
                    # 获取该用户组关联的所有策略
                    for policy_str in groups_to_policy[group]:
                        # policy = policy_str.strip('"')  # 去掉引号以匹配逻辑（如果需要）
                        policy = policy_str
                        if policy not in user_policy_dict[user]:
                            user_policy_dict[user].add(policy)

    # 将结果存入字典，使用集合转换为列表（如果需要保持顺序，可以用其他方法）
    result_dict = {user: sorted(list(policies)) for user, policies in user_policy_dict.items()}  # 排序以便于查看

    return result_dict


## format 优化结果函数-------------------
# 处理指定列函数，去重，合并单元格
def format_excel(file_path, sheet_name="用户-策略关系", target_column='C'):
    """
    处理Excel文件，并对指定列去重后合并单元格。

    参数:
    - file_path: Excel文件路径
    - sheet_name: 工作表名称
    - target_column: 要处理的列（默认是C列，可以指定为D列等）
    """
    # 加载Excel文件
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 初始化变量
    start_row = None
    current_key = None
    merged_ranges = []  # 用于记录需要合并的单元格范围
    result_dict = {}  # 用于记录每个A列值的去重后的目标列内容
    target_col_index = ord(target_column.upper()) - ord('A') + 1  # 将列字母转换为列索引

    # 遍历A列
    for row in range(1, ws.max_row + 1):
        a_value = ws[f"A{row}"].value
        target_value = ws[f"{target_column}{row}"].value

        # 如果A列值改变或到达最后一行，处理之前的块
        if a_value != current_key:
            if current_key is not None and start_row is not None:
                # 收集目标列的去重内容
                if start_row not in result_dict:
                    unique_values = set()
                    for r in range(start_row, row):
                        if ws[f"{target_column}{r}"].value:
                            for line in str(ws[f"{target_column}{r}"].value).splitlines():  # 按行分割目标列内容
                                unique_values.add(line.strip())
                    result_dict[start_row] = unique_values
                # 记录需要合并的范围
                merged_ranges.append((start_row, row - 1))

            # 更新当前A列值和起始行
            current_key = a_value
            start_row = row

    # 处理最后一个块（如果文件没有以新A列值结束）
    if current_key is not None and start_row is not None:
        if start_row not in result_dict:
            unique_values = set()
            for r in range(start_row, ws.max_row + 1):
                if ws[f"{target_column}{r}"].value:
                    for line in str(ws[f"{target_column}{r}"].value).splitlines():  # 按行分割目标列内容
                        unique_values.add(line.strip())
            result_dict[start_row] = unique_values
        merged_ranges.append((start_row, ws.max_row))

    # 写入结果并合并单元格
    for start, end in merged_ranges:
        # # 合并A列
        # ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)

        # 写入去重后的目标列内容到start行，并合并目标列
        unique_values = result_dict[start]
        ws[f"{target_column}{start}"] = "\n".join(sorted(unique_values))  # 排序后写入，按需
        ws.merge_cells(start_row=start, start_column=target_col_index, end_row=end, end_column=target_col_index)

        # 设置合并单元格的对齐方式（包括自动换行）
        for cell in ws[f"{target_column}{start}:{target_column}{end}"]:
            for c in cell:
                c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    # 保存文件
    wb.save(file_path)


# 合并A列单元格（用户名）
def merge_a_column_cells(file_path, sheet_name):
    """
    合并指定工作表中A列的连续相同值的单元格。

    参数:
    - file_path: Excel文件的路径
    - sheet_name: 工作表的名称
    """
    # 加载工作簿和工作表
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 初始化变量
    start_row = None
    current_key = None

    # 遍历A列
    for row in range(1, ws.max_row + 1):
        a_value = ws[f"A{row}"].value

        # 如果A列值改变或到达最后一行，处理之前的块
        if a_value != current_key:
            if current_key is not None and start_row is not None:
                # 合并单元格
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                # 可选：设置合并后单元格的对齐方式
                merged_cell = ws[f"A{start_row}"]
                merged_cell.alignment = Alignment(vertical="center", horizontal="left")

            # 更新当前A列值和起始行
            current_key = a_value
            start_row = row

    # 处理最后一个块（如果文件没有以新A列值结束）
    if current_key is not None and start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
        # 可选：设置合并后单元格的对齐方式
        merged_cell = ws[f"A{start_row}"]
        merged_cell.alignment = Alignment(vertical="center", horizontal="left")

    # 保存工作簿
    wb.save(file_path)

# 设置列宽、字体、背景色
def create_styled_excel(filename, sheet_name="用户-策略关系", header_fill_color="CCFFCC", font_size=12, column_width=30):
    # 加载excel
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    # 设置第一行的背景色、字体大小、加粗
    fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")  # 背景色
    font = Font(bold=True, size=font_size)  # 字体大小和加粗

    # 定义边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用样式到第一行
    for cell in ws[1]:  # ws[1] 表示第一行
        cell.fill = fill
        cell.font = font
        cell.border = thin_border

    # 应用边框到整个工作表的所有单元格
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1:
                # 第一行已经设置过样式，这里可以跳过或重新设置（如果需要）
                continue
            # 对于非第一行的单元格，仅设置边框（如果需要统一边框，也可以包括第一行）
            cell.border = thin_border

    # 设置所有列的宽度
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = column_width

    # 保存工作簿
    wb.save(filename)

# 删除sheet
def delete_sheet(output_file_path, sheet_name=["策略-目的IP关系", "策略-服务关系"]):
    workbook = load_workbook(output_file_path)
    for sheet in sheet_name:
        if sheet in workbook.sheetnames:
            del workbook[sheet]
    workbook.save(output_file_path)

# 工具
# 列表自动填写到单元格里
def write_list_to_excel(ws=None, data_list=None, row=1, col=1):
    """
    将列表内容写入Excel单元格，以换行符分隔，并设置单元格为自动换行。

    参数:
    ws (Worksheet): 现有工作表对象，如果提供，则不创建新工作簿。
    data_list (list): 要写入的列表数据。
    row (int): 起始行号。
    col (int): 起始列号。
    """
    # 将列表内容合并为字符串，以换行符分隔
    cell_value = "\n".join(map(str, data_list))
    # 写入数据到指定单元格
    target_cell = ws.cell(row=row, column=col)
    existing_content = target_cell.value
    if existing_content:
        target_cell.value = str(existing_content) + '\n' + cell_value
    else:
        target_cell.value = cell_value
    # 设置单元格为自动换行
    target_cell.alignment = Alignment(wrapText=True)
    return ws  # 可选：返回ws以便进一步操作
# 示例用法
if __name__ == "__main__":
    # 获取配置文件路径
    print("程序运行期间不要打开所导出的excel文件！！")
    print("请选择配置文件！")
    input_file_path, input_filename = select_file_and_get_path()
    ## 截取文件内容
    ### 获取用户配置信息
    user_text = extract_config_content(input_file_path, title="config user local")
    #### 获取用户列表
    user_list = get_user_list(text=user_text)

    ### 获取policy配置信息
    policy_text = extract_config_content(input_file_path, title="config firewall policy")
    #### 获取IP列表
    IP_list = extract_dstaddr_lines(text=policy_text, pattern=r'set\s+dstaddr\s+(.*)')
    #### 获取服务列表
    service_list = extract_dstaddr_lines(text=policy_text, pattern=r'set\s+service\s+(.*)')

    ##### 获取IP组文本
    ip_text = extract_config_content(input_file_path, title="config firewall address", title2="config firewall addrgrp")
    ##### 获取IP组-IP字典
    ip_dict = ipgroup_ip(ip_text)
    ##### 获取服务组-服务文本
    server_text = extract_config_content(input_file_path, title="config firewall service custom", title2="config firewall service custom")
    #### 获取字典



    #### 调用函数解析文本,获取数据字典
    group_policy_map, user_policy_map, dstaddr_dict, service_dict, status_dict, name_dict = parse_groups_and_users(policy_text)
    ##### 获取用户组配置信息
    group_text = extract_config_content(input_file_path, title="config user group")
    ### 调用函数解析文本，获取用户组-用户信息
    group_user_map = user_group(group_text)
    #### 获取用户组-用户字典
    user_policy_dict = find_user_policies(user_list, user_policy_map, group_policy_map, group_user_map)
    ##### 反转，获取用户-用户组字典
    user_group_map = rever_dict(group_user_map)




    # 写入excel，指定excel表格名称
    # 获取excel表格导出路径
    print("请选择导出 用户策略表.xlsx 存放文件夹！")
    output_file_path = select_folder_and_get_path(input_filename)
    print("导出文件：" + output_file_path)
    # file_path= "用户策略.xlsx"


    # 写入用户-策略关系
    print("写入用户策略关系中...")
    edit_excel_sheet(file_path=output_file_path, sheet_name="用户-策略关系", groups=user_policy_dict, A_name="用户", B_name="所绑定策略ID")

    # 写入用户组
    fill_excel_column(file_path=output_file_path, map=user_group_map,title="用户组", source_col="A", target_col="C")

    # 写入策略名称
    fill_excel_column(file_path=output_file_path, map=name_dict,title="策略名称", source_col="B", target_col="D")
    # 写入策略状态
    fill_excel_column(file_path=output_file_path, map=status_dict,title="策略状态", source_col="B", target_col="E")

    # 写入目的IP组
    fill_excel_column(file_path=output_file_path, map=dstaddr_dict,title="目的IP组", source_col="B", target_col="F")

    # 写入具体IP
    fill_excel_column(file_path=output_file_path, map=ip_dict, title="详细IP", source_col="F", target_col="G")

    # 写入服务
    fill_excel_column(file_path=output_file_path, map=service_dict,title="服务", source_col="B", target_col="H")

    # 正在格式化...
    print("格式化中...")
    # 合并单元格
    format_excel(file_path=output_file_path, target_column='C')  # 可以改为 'D' 或其他列
    # format_excel(file_path=output_file_path, target_column='D')  # 可以改为 'D' 或其他列
    merge_a_column_cells(file_path=output_file_path, sheet_name="用户-策略关系")

    # 边框、字体、列宽
    create_styled_excel(output_file_path, header_fill_color="CCFFCC", font_size=12, column_width=38)
    # 删除sheet
    delete_sheet(output_file_path=output_file_path)
    input("导出成功！\n回车键退出（或直接关闭程序）....")
